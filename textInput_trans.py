import time
import pygame, sys
from button import Button
from pygame.locals import *
from rule import *
from gomoku import Board, Gomoku
import numpy as np
from tensorflow.keras.models import load_model
import cv2
from goboard_edge_detect_new import goboard_edge_detect_module, centroids_sort, index_to_coordinate, data_stone_package
from gostone_matching_new import gostone_matching_module, stone_55_list
import pandas as pd
from openpyxl import load_workbook
import serial
from time import sleep

w, h = 19, 19
board_1 = Board(w=w, h=h)
game = Gomoku(board=board_1)
model = load_model("20221102_122102_1919_10.h5")
# model = load_model('20210307_232530.h5')
global centroids, flag, img_board, check_stone, black_stone_n, check_num_pg, first_turn
global white_stone_n, now_black_stone, black_stone_pos_x, black_stone_pos_y, cap, b, name_index

# py_serial = serial.Serial(port='COM4', baudrate=57600)

bg_color = (128, 128, 128)
black = (0, 0, 0)
blue = (0, 50, 255)
white = (255, 255, 255)
red = (255, 0, 0)
green = (0, 200, 0)

window_width = 1280
window_height = 800
board_width = 800
grid_size = 40

fps = 60
fps_clock = pygame.time.Clock()

pygame.init()

SCREEN = pygame.display.set_mode((1280, 800))
pygame.display.set_caption("Menu")

BG = pygame.image.load("image/background2.jpg")
stone_image = pygame.image.load("image/omokstone.png")
stone_image = pygame.transform.scale(stone_image, (240, 240))
robot_image = pygame.image.load("image/robot.png")
robot_image = pygame.transform.scale(robot_image, (150, 150))
all_image = pygame.image.load("image/all.png")
loading_image = pygame.image.load("image/loading.jpg")
# choice_image = pygame.image.load("image/choice.png")
font = pygame.font.Font(None, 32)
font2 = pygame.font.Font(None, 70)
data = load_workbook('ranking.xlsx')
ws = data.active
name_list = []

def get_font(size):  # Returns Press-Start-2P in the desired size
    return pygame.font.Font("tway_sky.ttf", size)


def play():

    SCREEN.blit(BG, (0, 0))

    OPTIONS_TEXT = get_font(70).render("Loading . . .", True, "Black")
    OPTIONS_TEXT_loc = OPTIONS_TEXT.get_rect(center=(640, 160))
    tip_image = pygame.image.load("image/Options Rect.png")
    tip_image = pygame.transform.scale(tip_image, (1200, 100))
    OPTIONS_tip = get_font(20).render("Tip! 돌을 착수하는 시간이 초과되면 패널티가 적용되고 3개의 목숨이 사라지면 게임에서 패배하게 됩니다!  ", True, "Green")
    OPTIONS_tip_loc = OPTIONS_TEXT.get_rect(center=(350, 730))
    SCREEN.blit(tip_image, (50, 650))
    SCREEN.blit(OPTIONS_TEXT, OPTIONS_TEXT_loc)
    SCREEN.blit(OPTIONS_tip, OPTIONS_tip_loc)

    SCREEN.blit(all_image, (1100, 10))

    pygame.display.update()
    pygame.init()
    surface = pygame.display.set_mode((window_width, window_height))
    pygame.display.set_caption("Omok game")
    surface.fill(bg_color)

    omok = Omok(surface)
    menu = Menu(surface)
    while True:
        run_game(surface, omok, menu)
        menu.is_continue(omok)


def ranking():
    data_rank = pd.read_excel('ranking.xlsx')
    data_rank = data_rank.sort_values(by=data_rank.columns[3], ascending=False)
    data_rank = data_rank.head(3)
    data_rank = data_rank.values.tolist()
    data_list = []
    for i in range(0, 3):
        data_list.append(' '.join(str(e) for e in data_rank[i][:]))
    text_1st = data_list[0]
    text_1st = text_1st.split()
    text_2nd = data_list[1]
    text_2nd = text_2nd.split()
    text_3rd = data_list[2]
    text_3rd = text_3rd.split()

    while True:
        font = pygame.font.Font(None, 32)
        rec = pygame.image.load("image/rank_bg.png")


        OPTIONS_MOUSE_POS = pygame.mouse.get_pos()

        SCREEN.blit(BG, (0, 0))
        SCREEN.blit(rec, (200, 0))

        txt_name = font.render("name", True, "white")
        txt_win = font.render("win", True, "white")
        txt_defeat = font.render("defeat", True, "white")
        txt_win_rate = font.render("win rate", True, "white")

        SCREEN.blit(txt_name, (420, 300))
        SCREEN.blit(txt_win, (590, 300))
        SCREEN.blit(txt_defeat, (730, 300))
        SCREEN.blit(txt_win_rate, (890, 300))

        for i in range(0, 4):
            if i == 0:
                txt_1st = font.render(text_1st[i], True, "white")
                SCREEN.blit(txt_1st, (400, 390))
            else:
                txt_1st = font.render(text_1st[i], True, "white")
                SCREEN.blit(txt_1st, (i * 150 + 450, 390))

        for i in range(0, 4):
            if i == 0:
                txt_2nd = font.render(text_2nd[i], True, "white")
                SCREEN.blit(txt_2nd, (400, 525))
            else:
                txt_2nd = font.render(text_2nd[i], True, "white")
                SCREEN.blit(txt_2nd, (i * 150 + 450, 525))

        for i in range(0, 4):
            if i == 0:
                txt_3rd = font.render(text_3rd[i], True, "white")
                SCREEN.blit(txt_3rd, (400, 650))
            else:
                txt_3rd = font.render(text_3rd[i], True, "white")
                SCREEN.blit(txt_3rd, (i * 150 + 450, 650))

        OPTIONS_BACK = Button(image=None, pos=(1200, 750),
                              text_input="BACK", font=get_font(30), base_color="white", hovering_color="Green")

        OPTIONS_BACK.changeColor(OPTIONS_MOUSE_POS)
        OPTIONS_BACK.update(SCREEN)

        for event in pygame.event.get():
            if event.type == pygame.QUIT:
                pygame.quit()
                sys.exit()
            if event.type == pygame.MOUSEBUTTONDOWN:
                if OPTIONS_BACK.checkForInput(OPTIONS_MOUSE_POS):
                    main_menu()

        pygame.display.update()



def choice_first():

    global first_turn

    while True:
        SCREEN.blit(loading_image, (0, 0))
        SCREEN.blit(all_image, (1100, 10))
        # SCREEN.blit(choice_image, (100, 20))
        choice_first_pos = pygame.mouse.get_pos()
        MENU_TEXT = font2.render("Choice your stone", True, "#ffffff")
        tip_image = pygame.image.load("image/Options Rect.png")
        tip_image = pygame.transform.scale(tip_image, (500, 100))
        MENU_RECT = MENU_TEXT.get_rect(center=(640, 100))
        SCREEN.blit(tip_image, (400, 50))
        SCREEN.blit(MENU_TEXT, MENU_RECT)

        black_button = Button(image=pygame.image.load("image/Options Rect.png"), pos=(640, 350),
                              text_input="Black (선공)", font=get_font(50), base_color="#d7fcd4",
                              hovering_color="White")
        white_button = Button(image=pygame.image.load("image/Options Rect.png"), pos=(640, 500),
                              text_input="white (후공)", font=get_font(50), base_color="#d7fcd4",
                              hovering_color="White")
        back_button = Button(image=pygame.image.load("image/Quit Rect.png"), pos=(640, 650),
                             text_input="back", font=get_font(50), base_color="#d7fcd4", hovering_color="White")

        for button in [black_button, white_button, back_button]:
            button.changeColor(choice_first_pos)
            button.update(SCREEN)

        for event in pygame.event.get():
            if event.type == pygame.QUIT:
                pygame.quit()
                sys.exit()
            if event.type == pygame.MOUSEBUTTONDOWN:
                if black_button.checkForInput(choice_first_pos):
                    first_turn = 1
                    play()
                if white_button.checkForInput(choice_first_pos):
                    first_turn = 2
                    play()
                if back_button.checkForInput(choice_first_pos):
                    main_menu()

        pygame.display.update()


def name_input():
    screen = pygame.display.set_mode((1280, 800))
    clock = pygame.time.Clock()
    input_box = pygame.Rect(560, 200, 400, 32)

    tip_image = pygame.image.load("image/Options Rect.png")
    tip_image = pygame.transform.scale(tip_image, (1200, 100))
    OPTIONS_tip = get_font(20).render("  이름은 영어로 입력해주세요! 또한 다른 사람의 이름을 도용하면 안 됩니다!  ", True, "Green")

    color_inactive = pygame.Color('white')
    color_active = pygame.Color('dodgerblue2')
    color = color_inactive
    rec = pygame.image.load("image/loading2.jpg")

    text = ''
    active = False
    done = False

    for row in ws.rows:
        name_list.append(row[0].value)

    while not done:
        for event in pygame.event.get():
            if event.type == pygame.QUIT:
                done = True
            if event.type == pygame.MOUSEBUTTONDOWN:
                # If the user clicked on the input_box rect.
                if input_box.collidepoint(event.pos):
                    # Toggle the active variable.
                    active = not active
                else:
                    active = False
                # Change the current color of the input box.
                color = color_active if active else color_inactive
            if event.type == pygame.KEYDOWN:
                if active:
                    if event.key == pygame.K_RETURN:
                        name = text
                        text = ''
                        if name in name_list:
                            new_name(1, name)
                        else:
                            new_name(2, name)
                    elif event.key == pygame.K_BACKSPACE:
                        text = text[:-1]
                    else:
                        text += event.unicode

        screen.fill((0, 0, 0))
        screen.blit(rec, (0, 0))
        screen.blit(all_image, (1100, 10))
        name_text = font.render("input your name and press \"enter\"", True, "#ffffff")
        text_name = font.render("Name : ", True, "#ffffff")
        pygame.draw.rect(screen, (255, 255, 255), (310, 20, 670, 750), 3)
        screen.blit(name_text, (450, 100))
        screen.blit(text_name, (450, 200))
        SCREEN.blit(tip_image, (50, 650))
        SCREEN.blit(OPTIONS_tip, (280, 690))

        # Render the current text.
        txt_surface = font.render(text, True, color)

        # Resize the box if the text is too long.
        width = max(200, txt_surface.get_width() + 10)
        input_box.w = width
        # Blit the text.
        screen.blit(txt_surface, (input_box.x + 5, input_box.y + 5))
        # Blit the input_box rect.
        pygame.draw.rect(screen, color, input_box, 2)

        pygame.display.flip()
        clock.tick(30)


def new_name(num, name):
    global name_index
    tip_image = pygame.image.load("image/Options Rect.png")
    tip_image = pygame.transform.scale(tip_image, (700, 100))

    while True:
        MENU_POS = pygame.mouse.get_pos()
        SCREEN.blit(BG, (0, 0))
        if num == 1:
            add_name = "next"
            main_text = get_font(30).render(f"Is your name is \"{name}\"?", True, "#ffffff")
            flag_name = False
        else:
            add_name = "add Name"
            main_text = get_font(30).render("your name is not in the database", True, "#ffffff")
            flag_name = True

        main_text_pos = main_text.get_rect(center=(640, 100))

        SCREEN.blit(tip_image, (300, 50))
        SCREEN.blit(main_text, main_text_pos)

        add_name_button = Button(image=pygame.image.load("image/Play Rect.png"), pos=(450, 500),
                                     text_input=add_name, font=get_font(40), base_color="#d7fcd4",
                                     hovering_color="White")

        back_button = Button(image=pygame.image.load("image/Play Rect.png"), pos=(850, 500),
                                 text_input="back", font=get_font(40), base_color="#d7fcd4", hovering_color="White")

        for button in [add_name_button, back_button]:
            button.changeColor(MENU_POS)
            button.update(SCREEN)

        for event in pygame.event.get():
            if event.type == pygame.QUIT:
                pygame.quit()
                sys.exit()
            if event.type == pygame.MOUSEBUTTONDOWN:
                if add_name_button.checkForInput(MENU_POS):
                    if flag_name == True:
                        ws["A" + str((len(name_list))+1)] = name
                        ws["B" + str((len(name_list))+1)] = 0
                        ws["C" + str((len(name_list))+1)] = 0
                        name_index = len(name_list)
                        flag_name = False
                        choice_first()
                    else:
                        name_index = name_list.index(name)
                        flag_name = False
                        choice_first()

                if back_button.checkForInput(MENU_POS):
                    main_menu()

        pygame.display.update()


def main_menu():

    while True:
        SCREEN.blit(BG, (0, 0))
        # SCREEN.blit(stone_image, (0, 560))
        SCREEN.blit(robot_image, (1130, 650))
        SCREEN.blit(all_image, (1100, 10))
        MENU_MOUSE_POS = pygame.mouse.get_pos()

        MENU_TEXT = get_font(100).render("Omok Game", True, "#000000")
        MENU_RECT = MENU_TEXT.get_rect(center=(640, 100))

        PLAY_BUTTON = Button(image=pygame.image.load("image/Play Rect.png"), pos=(640, 350),
                             text_input="Play", font=get_font(50), base_color="#d7fcd4", hovering_color="White")
        OPTIONS_BUTTON = Button(image=pygame.image.load("image/Options Rect.png"), pos=(640, 500),
                                text_input="Rankings", font=get_font(50), base_color="#d7fcd4", hovering_color="White")
        QUIT_BUTTON = Button(image=pygame.image.load("image/Quit Rect.png"), pos=(640, 650),
                             text_input="Quit", font=get_font(50), base_color="#d7fcd4", hovering_color="White")

        SCREEN.blit(MENU_TEXT, MENU_RECT)

        for button in [PLAY_BUTTON, OPTIONS_BUTTON, QUIT_BUTTON]:
            button.changeColor(MENU_MOUSE_POS)
            button.update(SCREEN)

        for event in pygame.event.get():
            if event.type == pygame.QUIT:
                pygame.quit()
                sys.exit()
            if event.type == pygame.MOUSEBUTTONDOWN:
                if PLAY_BUTTON.checkForInput(MENU_MOUSE_POS):
                    name_input()
                if OPTIONS_BUTTON.checkForInput(MENU_MOUSE_POS):
                    ranking()
                if QUIT_BUTTON.checkForInput(MENU_MOUSE_POS):
                    pygame.quit()
                    sys.exit()

        pygame.display.update()


def run_game(surface, omok, menu):
    global centroids, flag, img_board, check_stone, black_stone_n, check_num_pg, elapsed_time, timer, timer_flag
    global white_stone_n, now_black_stone, black_stone_pos_x, black_stone_pos_y, cap, b, first_turn, seq_flag, POS_white

    w, h = 19, 19

    seq_flag = 0
    timer_flag = 0


    size_of_board = 19
    board_array = np.zeros((size_of_board, size_of_board), dtype=np.int8)
    # board_array = [[0 for x in range(20)] for y in range(20)]
    print(board_array)
    black_player = 1
    white_player = 2
    game_result = 0

    max_turn = size_of_board * size_of_board

    black_stone_n = 0
    white_stone_n = 0

    centroids = []
    check_stone = []
    board_buttons = [[0 for x in range(19)] for y in range(19)]
    board_buttons_compare = [[0 for x in range(19)] for y in range(19)]

    all_white_stone = []
    all_black_stone = []

    cap = cv2.VideoCapture(1)
    flag = True
    total_time = 60

    start_ticks = pygame.time.get_ticks()

    omok.init_game()

    cap = cv2.VideoCapture(1)

    while True:

        check, frame = cap.read()
        img_realtime = frame.copy()
        img_realtime = cv2.resize(img_realtime, (640, 480), interpolation=cv2.INTER_CUBIC)
        cv2.imshow("img_realtime", img_realtime)
        key = cv2.waitKey(10)

        omok.surface.blit(omok.back_image, (800, 0))
        omok.surface.blit(omok.bsimage, (850, 50))
        omok.surface.blit(omok.wsimage, (850, 500))
        black_image = pygame.image.load("image/black.png")
        black_image = pygame.transform.scale(black_image, (grid_size, grid_size))
        white_image = pygame.image.load("image/white.png")
        white_image = pygame.transform.scale(white_image, (grid_size, grid_size))
        time_image = pygame.image.load("image/time_back.png")
        omok.surface.blit(time_image, (950, 0))
        elapsed_time = (pygame.time.get_ticks() - start_ticks) / 1000
        timer = pygame.font.Font("Digital.ttf", 50).render("time : " + str(int(total_time - elapsed_time)), True, (255, 255, 255))
        omok.surface.blit(timer, (950, 20))

        omok.surface.blit(omok.heart, (850, 50))
        omok.surface.blit(omok.heart, (950, 50))
        omok.surface.blit(omok.heart, (1050, 50))

        if total_time - elapsed_time <= 0:
            start_ticks = pygame.time.get_ticks()
            timer_flag += 1

        if timer_flag == 1:
            omok.surface.blit(omok.heart_g, (862, 65))
        elif timer_flag == 2:
            omok.surface.blit(omok.heart_g, (862, 65))
            omok.surface.blit(omok.heart_g, (962, 65))
        elif timer_flag == 3:
            omok.surface.blit(omok.heart_g, (862, 65))
            omok.surface.blit(omok.heart_g, (962, 65))
            omok.surface.blit(omok.heart_g, (1062, 65))
            omok.surface.blit(omok.defeat, (400, 300))
            omok.show_end_menu()

        pygame.display.update()
        fps_clock.tick(fps)

        if check and key == ord('q'):
            img = frame.copy()
            dst, img_rgb, stone_list = gostone_matching_module(img, first_turn)
            stone_coordinate_list = stone_55_list(stone_list)

            if len(centroids) != 361:
                img_board, centroids = goboard_edge_detect_module(img)
            else:
                if flag == True:
                    b = centroids_sort(centroids)
                    flag = False

                for index, pt in enumerate(b):
                    cv2.putText(img_board, str(index), (int(pt[0]), int(pt[1])), cv2.FONT_HERSHEY_DUPLEX, 0.3,
                                (0, 255, 0))
                    cv2.circle(img_board, (int(pt[0]), int(pt[1])), 3, (0, 0, 255))
                    check1 = list([int(pt[0]), int(pt[1])])

                    if len(stone_coordinate_list) != len(check_stone):
                        for stone in stone_coordinate_list:
                            if check1 in stone:
                                x, y = index_to_coordinate(index)
                                board_buttons[x][y] = 1
                                check_stone.append((x, y))
                                check_stone = list(set(check_stone))

                for i in range(19):
                    for j in range(19):
                        if board_buttons[i][j] != board_buttons_compare[i][j]:
                            now_black_stone = ([i, j])
                            all_black_stone.append(now_black_stone)
                            print("all_black_stone =", all_black_stone)
                            pos_H, pos_W = i, j

                            board_array[pos_H, pos_W] = 1
                            black_stone_n = black_stone_n + 1
                            # game_result = game_rule(board_array, black_player)

                            data_output_black = data_stone_package(i, j)
                            print("현재 착수된 검은돌의 위치", now_black_stone, data_output_black)
                            board_buttons_compare[i][j] = board_buttons[i][j]

                if first_turn == 2:
                    omok.turn = 1
                    first_black = (9 * grid_size + 40, 9 * grid_size + 40)
                    omok.check_board(first_black)
                    board_array[9, 9] = 2
                    omok.turn = 2

                if black_stone_n > white_stone_n:
                    if board_buttons == board_buttons_compare:
                        input_1 = board_array.copy()
                        input_1[(input_1 != 1) & (input_1 != 0)] = -1
                        input_1[(input_1 == 1) & (input_1 != 0)] = 1
                        input_1 = np.expand_dims(input_1, axis=(0, -1)).astype(np.float32)

                        output = model.predict(input_1).squeeze()
                        output = output.reshape((h, w))
                        output_y, output_x = np.unravel_index(np.argmax(output), output.shape)
                        data_output_white = data_stone_package(output_y, output_x)
                        now_white_stone = ([output_y, output_x])
                        all_white_stone.append(now_white_stone)
                        print("all_white_stone =", all_white_stone)
                        print("현재 착수된 백돌의 위치", now_white_stone, data_output_white)

                        pos_H, pos_W = int(output_x), int(output_y)
                        board_array[pos_W, pos_H] = 2  # black is one
                        white_stone_n = white_stone_n + 1

                        black_stone_pos_x = now_black_stone[0]
                        black_stone_pos_y = now_black_stone[1]
                        print(black_stone_pos_x)
                        print(black_stone_pos_y)
                        trans_b = black_stone_pos_x
                        black_stone_pos_x = (18 - black_stone_pos_y)
                        black_stone_pos_y = trans_b
                        print(black_stone_pos_x)
                        print(black_stone_pos_y)
                        black_stone_pos = (black_stone_pos_y * grid_size + 40, black_stone_pos_x * grid_size + 40)
                        print(black_stone_pos)
                        omok.check_board(black_stone_pos)
                        seq_flag += 1
                        if first_turn == 1:
                            if seq_flag == 1:
                                print(first_turn)
                            else:
                                omok.surface.blit(white_image, (POS_white[0]-20, POS_white[1]-20))
                        else:
                            if seq_flag == 1:
                                omok.surface.blit(black_image, (9 * grid_size + 20, 9 * grid_size + 20))
                            else:
                                omok.surface.blit(black_image, (POS_white[0]-20, POS_white[1]-20))

                        if omok.is_gameover:
                            return

                        pygame.display.update()
                        fps_clock.tick(fps)
                        check_num_pg = 2

                        if check_num_pg == 2:
                            time.sleep(1.0)
                            output_x_w = now_white_stone[0]
                            output_y_w = now_white_stone[1]
                            trans_w = output_x_w
                            output_x_w = 18 - output_y_w
                            output_y_w = trans_w
                            POS_white = [output_y_w * grid_size + 40, output_x_w * grid_size + 40]
                            print(POS_white)
                            # commend = str(data_output_white)
                            # py_serial.write(commend.encode())
                            omok.check_board(POS_white)

                            seq_flag += 1
                            if first_turn == 1:
                                if seq_flag == 2:
                                    omok.surface.blit(black_image, (black_stone_pos[0]-20, black_stone_pos[1]-20))
                                else:
                                    omok.surface.blit(black_image, (black_stone_pos[0]-20, black_stone_pos[1]-20))
                            else:
                                if seq_flag == 2:
                                    omok.surface.blit(white_image, (black_stone_pos[0] - 20, black_stone_pos[1] - 20))
                                else:
                                    omok.surface.blit(white_image, (black_stone_pos[0]-20, black_stone_pos[1]-20))

                            start_ticks = pygame.time.get_ticks()

                            if omok.is_gameover:
                                return

                            pygame.display.update()
                            fps_clock.tick(fps)

            print(board_array)

            cv2.imshow("img_board", img_board)
            cv2.imshow("img_rgb", img_rgb)

            pygame.display.update()
            fps_clock.tick(fps)

        if cv2.waitKey(1) & 0xFF == 27:  # esc 키를 누르면 닫음
            break
    if seq_flag == 400:
        seq_flag = 0

    cap.release()
    cv2.destroyAllWindows()


class Omok(object):
    def __init__(self, surface):
        self.turn = 1
        self.board = [[0 for i in range(board_size)] for j in range(board_size)]
        self.menu = Menu(surface)
        self.rule = Rule(self.board)
        self.surface = surface
        self.pixel_coords = []
        self.set_coords()
        self.set_image_font()
        self.is_show = True

    def init_game(self):
        if first_turn == 1:
            self.turn = 1
        elif first_turn == 2:
            self.turn = 2
        self.draw_board()
        # self.menu.show_msg(empty)
        self.init_board()
        self.coords = []
        self.redos = []
        self.id = 1
        self.is_gameover = False
        self.is_forbidden = False

    def set_image_font(self):
        black_img = pygame.image.load('image/black_tri.png')
        white_img = pygame.image.load('image/white_tri.png')
        self.last_w_img = pygame.image.load('image/white_a.png')
        self.last_b_img = pygame.image.load('image/black_a.png')
        self.board_img = pygame.image.load('image/board_last1.png')
        self.omokBoard_img = pygame.image.load('image/board_last1.png')
        self.back_image = pygame.image.load("image/floor.jpg")
        self.bsimage = pygame.image.load("image/white_st.png")
        self.wsimage = pygame.image.load("image/black_st.png")
        self.win = pygame.image.load("image/win2.png")
        self.defeat = pygame.image.load("image/defeat.png")
        self.heart = pygame.image.load("image/heart.png")
        self.heart_g = pygame.image.load("image/heart_g.png")
        self.forbidden_img = pygame.image.load('image/forbidden.png')
        self.font = pygame.font.Font("freesansbold.ttf", 14)
        self.black_img = pygame.transform.scale(black_img, (grid_size, grid_size))
        self.white_img = pygame.transform.scale(white_img, (grid_size, grid_size))

    def init_board(self):
        for y in range(board_size):
            for x in range(board_size):
                self.board[y][x] = 0

    def draw_board(self):
        self.surface.blit(self.omokBoard_img, (0, 0))
        self.surface.blit(self.back_image, (800, 0))
        self.surface.blit(self.bsimage, (850, 50))
        self.surface.blit(self.wsimage, (850, 500))

    def draw_image(self, img_index, x, y):
        img = [self.black_img, self.white_img]
        self.surface.blit(img[img_index-1], (x, y))

    def draw_stone(self, coord, stone, increase):
        for i in range(len(self.coords)):
            x, y = self.coords[i]
            # self.draw_image(0, x-20, y-20)
        if self.coords:
            x, y = self.coords[-1]
            self.draw_image(self.turn, x - 20, y - 20)
        x, y = self.get_point(coord)
        self.board[y][x] = stone
        self.id += increase
        self.turn = 3 - self.turn

    def set_coords(self):
        for y in range(board_size):
            for x in range(board_size):
                self.pixel_coords.append((x * grid_size, y * grid_size))

    def get_coord(self, pos):
        for coord in self.pixel_coords:
            x, y = coord
            rect = pygame.Rect(x, y, grid_size, grid_size)
            if rect.collidepoint(pos):
                return coord
        return None

    def get_point(self, coord):
        x, y = coord
        x = (x - 25) // grid_size
        y = (y - 25) // grid_size
        return x, y

    def check_board(self, pos):
        coord = self.get_coord(pos)
        if not coord:
            return False
        x, y = self.get_point(coord)
        if self.board[y][x] != empty:
            print("occupied")
            return True

        # if self.turn == black_stone:
        #     if self.rule.forbidden_point(x, y, self.turn):
        #         print("forbidden point")
        #         return True

        self.coords.append(coord)
        self.draw_stone(coord, self.turn, 1)
        if self.check_gameover(coord, 3 - self.turn):
            self.is_gameover = True
        if len(self.redos):
            self.redos = []
        return True

    def check_gameover(self, coord, stone):
        x, y = self.get_point(coord)
        if self.id > board_size * board_size:
            self.show_winner_msg(stone)
            return True
        elif self.rule.is_gameover(x, y, stone):
            self.show_winner_msg(stone)
            return True
        return False

    def show_winner_msg(self, stone):
        #TODO 여기부터 시작
        global name_index
        if first_turn == 1:
            if stone == 1:
                ws["B" + str((name_index + 1))] = (int(ws["B" + str((name_index + 1))].value) + 1)
                data.save("ranking.xlsx")
                self.surface.blit(self.win, (400, 300))
                self.show_end_menu()

            if stone == 2:
                ws["C" + str((name_index + 1))] = (int(ws["C" + str((name_index + 1))].value) + 1)
                data.save("ranking.xlsx")
                self.surface.blit(self.defeat, (400, 300))
                self.show_end_menu()

        elif first_turn == 2:
            if stone == 1:
                ws["C" + str((name_index + 1))] = (int(ws["C" + str((name_index + 1))].value) + 1)
                data.save("ranking.xlsx")
                self.surface.blit(self.defeat, (400, 300))
                self.show_end_menu()

            if stone == 2:
                ws["B" + str((name_index + 1))] = (int(ws["B" + str((name_index + 1))].value) + 1)
                data.save("ranking.xlsx")
                self.surface.blit(self.win, (400, 300))
                self.show_end_menu()

    def show_end_menu(self):

        while True:
            menu_pos = pygame.mouse.get_pos()
            new_game_button = Button(image=pygame.image.load("image/Play Rect.png"), pos=(350, 500),
                                     text_input="New Game", font=get_font(40), base_color="#d7fcd4", hovering_color="White")

            main_button = Button(image=pygame.image.load("image/Play Rect.png"), pos=(750, 500),
                                 text_input="menu", font=get_font(40), base_color="#d7fcd4", hovering_color="White")

            for button in [new_game_button, main_button]:
                button.changeColor(menu_pos)
                button.update(self.surface)

            for event in pygame.event.get():
                if event.type == pygame.QUIT:
                    pygame.quit()
                    sys.exit()
                if event.type == pygame.MOUSEBUTTONDOWN:
                    if new_game_button.checkForInput(menu_pos):
                        self.new_game()
                    if main_button.checkForInput(menu_pos):
                        main_menu()

            pygame.display.update()

    def new_game(self):
        self.draw_board()
        self.init_game()
        play()

class Menu(object):
    def __init__(self, surface):
        self.font = pygame.font.Font('tway_sky.ttf', 30)
        self.surface = surface
        self.draw_menu()

    def draw_menu(self):
        top, left = window_height - 30, window_width - 200
        self.quit_rect = self.make_text(self.font, 'Quit Game', black, None, top, left)

    def make_text(self, font, text, color, bgcolor, top, left, position=0):
        surf = font.render(text, False, color, bgcolor)
        rect = surf.get_rect()
        if position:
            rect.center = (left, top)
        else:
            rect.topleft = (left, top)
        self.surface.blit(surf, rect)
        return rect

    def is_continue(self, omok):
        while True:
            for event in pygame.event.get():
                if event.type == QUIT:
                    self.terminate()

            pygame.display.update()
            fps_clock.tick(fps)

# def new_game():

if __name__ == '__main__':
    main_menu()
